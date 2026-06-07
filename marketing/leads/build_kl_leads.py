#!/usr/bin/env python3
"""Build the Elevare B2B outreach workbook — Kuala Lumpur tab, sector-categorized.
Kept separate from Chiang Mai (own tab) so the two lists never mix.
Data researched 2026-06-07 via web. Emails are ONLY real published addresses;
"FORM: <url>" means no public email — use their contact page instead. No invented contacts."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- brand colors ----
NAVY = "1B2A4A"
GOLD = "C9A227"
ROSE = "C2476B"
CREAM = "FBF7F0"
LIGHT = "EEF1F6"
WHITE = "FFFFFF"

SECTORS = [
    ("HEALTHCARE / MEDICAL", "Med-tier fit — staff treat foreign patients; OET / medical English"),
    ("HOSPITALITY / TOURISM", "General-tier fit — front-line staff serve foreign guests daily"),
    ("PROFESSIONAL SERVICES", "Pro-tier fit — staff write/negotiate with multinational clients"),
    ("CORPORATE / EXPORT / TECH", "Bulk buyer — teams communicate with international clients/partners"),
]

# columns: name, subsector, intl_angle, website, email, phone, location, source, confidence
healthcare = [
  ["Prince Court Medical Centre","Private hospital","MSQH-accredited medical-tourism hospital; patients from 140+ countries","https://princecourt.com","FORM: https://princecourt.com/internationalpatients","+603 2160 0000","Jalan Kia Peng, KL","https://princecourt.com/internationalpatients","verified"],
  ["Subang Jaya Medical Centre (SJMC)","Private hospital","JCI/ACHSI-accredited; International Patient Care & interpreter services","https://subangjayamedicalcentre.com","FORM: https://subangjayamedicalcentre.com","+603 5639 1212","Subang Jaya, Selangor","https://subangjayamedicalcentre.com/","verified"],
  ["Gleneagles Hospital Kuala Lumpur","Private hospital","Early JCI-accredited IHH hospital; international patient portal","https://www.gleneagles.com.my/kuala-lumpur","FORM: https://www.gleneagles.com.my/kuala-lumpur/contactus","+603 4141 3000","Jalan Ampang, KL","https://www.gleneagles.com.my/kuala-lumpur","verified"],
  ["Pantai Hospital Kuala Lumpur","Private hospital","IHH hospital with dedicated International Patient Centre","https://www.pantai.com.my/kuala-lumpur","my.phkl.international@pantai.com.my","+603 2296 0888","Jalan Bukit Pantai, KL","https://www.pantai.com.my/kuala-lumpur/contact-us","verified"],
  ["Sunway Medical Centre","Private hospital","JCI/ACHS-accredited; International Patient Centre, patients from 170+ countries","https://www.sunwaymedical.com","FORM: https://www.sunwaymedical.com/en/international-patient-centre","+6019-200 9191","Bandar Sunway, PJ","https://www.sunwaymedical.com/en/international-patient-centre","verified"],
  ["KPJ KL Dental Specialist Centre","Dental specialist centre","Health Tourism Dept with visa-extension support for foreign patients","https://www.kpjhealth.com.my/dental/health-tourism","info@kpjhealthcentre.com","+603 4021 0852","Kuala Lumpur","https://www.kpjhealth.com.my/dental/health-tourism","verified"],
  ["Beverly Wilshire Medical Centre","Aesthetic / cosmetic surgery","Cosmetic-vacation packages for intl medical-tourism clients","https://beverlywilshiremedical.com","FORM: https://beverlywilshiremedical.com/contact-us-kl/","+603 2118 2888","Kuala Lumpur","https://beverlywilshiremedical.com/","verified"],
  ["Beverly Wilshire Dental","Dental clinic","~40% intl clientele (AU/NZ/UK/Middle East) for dental tourism","https://www.beverlywilshiredental.com","FORM: https://www.beverlywilshiredental.com/contact-us/","+603 2118 2999","Kenanga Tower, Jln Tun Razak, KL","https://www.beverlywilshiredental.com/dental-tourism/","verified"],
  ["Da Vinci Clinic","Aesthetic / skin clinic","Doctor-led aesthetic clinic, multi-branch, regional clients","https://davinciclinic.com.my","FORM: https://davinciclinic.com.my/book-now/","+6010-283 2688","Mid Valley / Bukit Jalil / TRX, KL","https://davinciclinic.com.my/","verified"],
  ["Her Clinic","Aesthetic / skin clinic","LCP-certified, FDA-approved procedures, multi-branch KL/PJ","https://herclinic.my","admin@herclinic.my","+60 12-883 2503","Ampang, KL (+ Damansara, PJ)","https://herclinic.my/","verified"],
  ["RJ Clinic","Aesthetic / skin clinic","Serves overseas clients with multilingual support","https://www.rjclinic.my","care@rjclinic.my","+60 18-575 1353","Ativo Damansara, KL","https://www.rjclinic.my/","verified"],
  ["GLOJAS Specialist Clinic","Plastic surgery / hair restoration","Cosmetic specialist serving local + international patients","https://glojasaesthetic.com","info@glojasaesthetic.com","+603 6211 5555","Desa Sri Hartamas, KL","https://glojasaesthetic.com/","verified"],
  ["GEM Clinic","Aesthetic clinic","Premium aesthetic clinic in intl-visitor Mid Valley area","https://gem.clinic","FORM: https://gem.clinic/contact/","+6012-200 7035","Mid Valley City, KL","https://gem.clinic/","verified"],
  ["KL Fertility Centre","Fertility / IVF","Internationally trained specialists; patients 'from Malaysia and beyond'","https://www.klfertility.com","FORM: https://www.klfertility.com/contact-us/","+603 2780 4288","Bukit Damansara, KL","https://www.klfertility.com/contact-us/","verified"],
  ["Sunway Fertility Centre","Fertility / IVF","Dedicated intl patient support: travel, translation, transfer","https://sunwayfertility.com.my","FORM: https://sunwayfertility.com.my/contacts/","+6019-281 2337","Bandar Sunway, PJ","https://sunwayfertility.com.my/contacts/","verified"],
  ["Alpha Fertility Centre","Fertility / IVF","IVF for intl patients; multilingual support + ground arrangements","https://www.alphafertilitycentre.com","enquiry@alphafertilitycentre.com","+603 6141 6166","Kota Damansara, PJ","https://www.alphafertilitycentre.com/ivf-for-international-patient","verified"],
  ["Sunfert International Fertility Centre","Fertility / IVF","International fertility centre serving regional patients","https://www.sunfert.com","FORM: https://www.sunfert.com/clinics","+603 7622 8688","Bangsar South, KL","https://www.sunfert.com/clinics","verified"],
  ["ISEC (Intl Specialist Eye Centre)","Eye / ophthalmology","Tertiary eye centre; local, regional & international patients","https://www.isec.my","enquiries@isec.my","+603 2284 8989","Mid Valley, KL","https://www.isec.my/contact-us/","verified"],
  ["Dentalist","Dental clinic","Dental-tourism clinic serving international patients","https://dentalist.com.my","enquiry@dentalist.com.my","+6018-582 8587","North Kiara, Segambut, KL","https://dentalist.com.my/dental-tourism/","verified"],
  ["Imperial Dental Specialist Centre","Dental specialist centre","Patients from US/Canada/UK for dental tourism","https://imperialdsc.com","enquiries@imperialdsc.com","+603 9212 0605","Bangsar Baru, KL","https://imperialdsc.com/contact/","verified"],
  ["VISTA Eye Specialist","Eye / ophthalmology","Intl patient centre assists foreign patients with hotel selection","https://www.vista.com.my","FORM: https://www.vista.com.my/locations/","1 800 88 3937","The Curve, Mutiara Damansara, PJ","https://www.vista.com.my/","verified"],
  ["Institut Jantung Negara (IJN)","Cardiac specialist hospital","Intl Patient Centre with coordination, language & travel support","https://www.ijn.com.my/","international@ijn.com.my","+603 2600 6336","145 Jalan Tun Razak, KL","https://www.ijn.com.my/contact-us","listed"],
  ["Beacon Hospital","Oncology / cancer specialist hospital","Intl Services Dept; multilingual staff for overseas cancer patients","https://www.beaconhospital.com.my/","info@beaconhospital.com.my","+603 7620 7979","Seksyen 51, Petaling Jaya","https://www.beaconhospital.com.my/contact-us","verified"],
  ["Cardiac Vascular Sentral KL (CVSKL)","Cardiac & vascular specialist hospital","Opposite KLIA Ekspres at KL Sentral for inbound medical travelers","https://www.cvskl.com/","info@cvskl.com","+603 2276 7000","KL Sentral, KL","https://www.cvskl.com/contact-us","verified"],
  ["Tung Shin Hospital","Private general hospital","Central Pudu hospital w/ English patient services, near tourist district","https://www.tungshin.com.my/","enquiry@tungshin.com.my","+603 2037 2288","102 Jalan Pudu, KL","https://www.tungshin.com.my/patients-and-visitors/talk-to-us/contact-us","verified"],
  ["Assunta Hospital","Private general hospital","Established PJ hospital serving Klang Valley expat community","https://assunta.com.my/","enquiries@assunta.com.my","+603 7872 3000","Jalan Templer, Petaling Jaya","https://assunta.com.my/about/contact","listed"],
  ["ParkCity Medical Centre","Private general hospital","JCI-accredited; intl patient team in expat-heavy Desa ParkCity","https://parkcitymedicalcentre.com/","FORM: https://parkcitymedicalcentre.com/enquiry","+603 5639 1616","Desa ParkCity, KL","https://parkcitymedicalcentre.com/contact-us","verified"],
  ["Thomson Hospital Kota Damansara","Multi-specialty hospital","Intl Patient Liaison Officer; interpreter + travel support","https://www.thomsonhospitals.com/","enquiries@tmclife.com","+603 6287 1111","Kota Damansara, Petaling Jaya","https://www.thomsonhospitals.com/contact-details","verified"],
  ["Sunway Medical Centre Velocity","Private general hospital","Intl Patient Centre; <4km from KLCC, coordinates care pre-arrival","https://www.sunwaymedicalvelocity.com.my/","FORM: https://www.sunwaymedicalvelocity.com.my/en/international-patient-centre/","+603 9772 9301","Sunway Velocity, KL","https://www.sunwaymedicalvelocity.com.my/en/international-patient-centre/","verified"],
  ["KPJ Tawakkal KL Specialist Hospital","Private specialist hospital","Intl Patient Centre (since 2010) within KPJ network","https://www.kpjhealth.com.my/tawakkal","tawakkal@kpjtawakkal.com","+603 4026 7777","Pekeliling, KL","https://www.kpjhealth.com.my/tawakkal","listed"],
  ["Damansara Specialist Hospital 2 (KPJ)","Private specialist hospital","Intl Patient Centre: airport transfers, teleconsult follow-up","https://www.kpjhealth.com.my/damansara2","FORM: https://www.kpjhealth.com.my/damansara2/international-patient-centre","+603 7717 3000","Damansara, Petaling Jaya","https://www.kpjhealth.com.my/damansara2/international-patient-centre","verified"],
  ["ALTY Orthopaedic Hospital","Orthopaedic specialist hospital","Ortho/spine hospital on embassy-belt Jalan Ampang; intl enquiries via WhatsApp","https://www.altyortho.com/","info@altyortho.com","+603 2787 0500","187 Jalan Ampang, KL","https://www.altyortho.com/contact-us","verified"],
  ["Cengild G.I. Medical Centre","Gastroenterology / GI & liver centre","GI/liver centre in intl business district Bangsar South","https://cengild.com/","FORM: https://cengild.com/contact-us/","+603 2242 7000","Nexus, Bangsar South, KL","https://cengild.com/contact-us/","verified"],
  ["Tun Hussein Onn National Eye Hospital (THONEH)","Eye specialist hospital","Intl Patients team: appointments, airport reception, transport, translation","https://thoneh.my/","enquiry@thoneh.com","+603 7718 1488","PJS 52, Petaling Jaya","https://thoneh.my/international-patients","verified"],
  ["Premier Clinic","Aesthetic / medical beauty clinic","24/7 English/BM/Chinese support; expat branches Bangsar/KL/TTDI","https://premier-clinic.com/","contactus@premier-clinic.com","+60 12 662 5552","Bangsar Baru, KL","https://premier-clinic.com/","listed"],
  ["Nexus Clinic","Aesthetic / anti-aging clinic","KL Golden Triangle; non-surgical aesthetic care for expats since 2001","https://www.nexus-clinic.com/","FORM: https://www.nexus-clinic.com/contact-us/","+603 2163 5699","Wisma UOA II, Jalan Pinang, KL","https://www.nexus-clinic.com/contact-us/","listed"],
  ["Klinik Dr Inder","Aesthetic / hair transplant clinic","30+ yrs; markets to AU/US patients, Chinese & Tamil pages","https://www.klinikdrinder.com/","FORM: https://www.klinikdrinder.com/contact-us/","+603 7932 1818","Seksyen 13, Petaling Jaya","https://www.klinikdrinder.com/contact-us/","verified"],
  ["Thomson Fertility Centre (TMC Fertility)","Fertility / IVF centre","Intl patient services; tailor-made travel plans, RTAC-accredited","https://thomsonfertility.com.my/","enquiries@tmclife.com","+603 6258 0000","Metro Prima, Kepong, KL","https://thomsonfertility.com.my/","listed"],
  ["Smile Avenue Dental Surgery (Publika)","Dental clinic (cosmetic / implant)","English-speaking, expat-heavy Publika; overseas-visitor reviews","https://www.smileavenuemalaysia.com/","FORM: https://www.smileavenuemalaysia.com/","+603 3005 4109","Publika, Solaris Dutamas, KL","https://www.smileavenuemalaysia.com/","verified"],
]

hospitality = [
  ["InterContinental Kuala Lumpur","5-star international hotel","Global IHG hotel hosting intl business & leisure travelers, KLCC","https://kualalumpur.intercontinental.com/","intercontinental.kualalumpur@ihg.com","+603 2782 6000","165 Jalan Ampang, KL","https://kualalumpur.intercontinental.com/contact/","verified"],
  ["EQ Kuala Lumpur","5-star luxury hotel","440-room luxury hotel near KLCC serving intl guests","https://www.eqkualalumpur.equatorial.com/","info@kul.equatorial.com","+603 2789 7777","Jalan Sultan Ismail, KL","https://www.eqkualalumpur.equatorial.com/","verified"],
  ["Shangri-La Kuala Lumpur","5-star luxury hotel","Iconic city hotel; world-class hospitality to intl guests","https://www.shangri-la.com/kualalumpur/shangrila/","slkl@shangri-la.com","+603 2032 2388","11 Jalan Sultan Ismail, KL","https://www.shangri-la.com/kualalumpur/shangrila/","listed"],
  ["Mandarin Oriental, Kuala Lumpur","5-star luxury hotel","Petronas-adjacent luxury hotel hosting intl travelers","https://www.mandarinoriental.com/en/kuala-lumpur/petronas-towers","mokul-reservations@mohg.com","+603 2380 8888","KLCC, KL","https://www.mandarinoriental.com/en/kuala-lumpur/petronas-towers/contact-us","listed"],
  ["The Ritz-Carlton, Kuala Lumpur","5-star luxury hotel","Golden Triangle luxury hotel serving guests worldwide","https://www.ritzcarlton.com/en/hotels/kulrz-the-ritz-carlton-kuala-lumpur/overview/","reservation@ritzcarltonkl.com","+603 2142 8000","168 Jalan Imbi, KL","https://www.ritzcarlton.com/en/hotels/kulrz-the-ritz-carlton-kuala-lumpur/","listed"],
  ["Grand Millennium Kuala Lumpur","5-star hotel","Bukit Bintang hotel catering to intl tourists & shoppers","https://www.millenniumhotels.com/en/kuala-lumpur/grand-millennium-hotel-kuala-lumpur/","FORM: https://www.millenniumhotels.com/en/kuala-lumpur/grand-millennium-hotel-kuala-lumpur/","","Bukit Bintang, KL","https://www.millenniumhotels.com/en/kuala-lumpur/grand-millennium-hotel-kuala-lumpur/","listed"],
  ["Pan Pacific Serviced Suites KL","Serviced apartments (expat)","Bukit Bintang suites for expat / intl long-stay guests","https://www.panpacific.com/en/serviced-suites/pp-ss-kuala-lumpur.html","enquiry.ppskul@panpacific.com","+603 2706 8688","Jalan Sultan Ismail, KL","https://www.panpacific.com/en/serviced-suites/pp-ss-kuala-lumpur.html","verified"],
  ["Somerset Kuala Lumpur (Ascott)","Serviced residences (expat)","Embassy Row residence for expat & intl business stays","https://www.discoverasr.com/en/somerset-serviced-residence/malaysia/somerset-kuala-lumpur","enquiry.kualalumpur@the-ascott.com","+603 2723 8888","187 Jalan Ampang, KL","https://www.discoverasr.com/en/somerset-serviced-residence/malaysia/somerset-kuala-lumpur","verified"],
  ["The Maple Suite","Serviced apartments (expat)","Expat accommodation for business & leisure since 1997","https://www.themaplesuite.com/","FORM: https://www.themaplesuite.com/","","Changkat Raja Chulan, KL","https://www.themaplesuite.com/","listed"],
  ["World Express (M) Sdn Bhd","Inbound tour operator / DMC","Malaysia's oldest DMC; inbound travel for foreign visitors","https://worldexpress.travel/","wxpkul@worldexpress.travel","+603 2148 9601","86 Jalan Raja Chulan, KL","https://worldexpress.travel/","listed"],
  ["GMTC (Global Mgmt Travel & Conf.)","DMC / MICE event management","End-to-end DMC & MICE for intl corporate groups","https://gmtc.com.my/","info@gmtc.com.my","+603 2037 9585","Kuala Lumpur","https://gmtc.com.my/contact-us/","verified"],
  ["MITRA Tours & Travel","Inbound tour operator / incentive","Inbound & incentive group travel for foreign tour groups","https://www.mitra.travel/","info@mitra.travel","+603 2779 1313","Wisma Central, Jalan Ampang, KL","https://www.mitra.travel/","verified"],
  ["HTC Travel Services (M)","Inbound tour operator","Award-winning inbound operator; own coach fleet for visitors","https://htctravel.com.my/","FORM: https://htctravel.com.my/","+603 9222 2255","Fraser Business Park, Jln Metro Pudu, KL","https://htctravel.com.my/","verified"],
  ["Star Travel (M) Sdn Bhd","Travel agency / tour operator","Personalised journeys globally; intl leisure & business travel","https://www.startravel.com.my/","FORM: https://www.startravel.com.my/","+603 2786 7555","42-44 Jalan Raja Abdullah, KL","https://www.startravel.com.my/","verified"],
  ["Embassy Alliance Travel Group","Inbound tour operator / DMC","Handles inbound Malaysia travel for international clients","https://embassyalliance.com/","FORM: https://embassyalliance.com/contact","","Kuala Lumpur","https://embassyalliance.com/tour-operator-in-malaysia/","listed"],
  ["KLGCC Convention Centre","MICE / convention venue","Hosts international conferences, exhibitions & corporate events","https://www.simedarbyproperty.com/hospitality-leisure/klgcc-convention-centre/","conventioncentre@simedarbyproperty.com","+603 2642 9088","1A Jalan Bukit Kiara 1, KL","https://www.simedarbyproperty.com/hospitality-leisure/klgcc-convention-centre/","listed"],
]

professional = [
  ["Azmi & Associates","Law firm","Cross-border M&A; listed in intl law directories","https://www.azmilaw.com/","general@azmilaw.com","+603 2118 5000","Menara Keck Seng, Jln Bukit Bintang, KL","https://www.azmilaw.com/contact/","verified"],
  ["Shearn Delamore & Co","Law firm","Full-service firm handling complex cross-border matters","https://www.shearndelamore.com/","info@shearndelamore.com","+603 2027 2727","Leboh Ampang, KL","https://www.shearndelamore.com/","verified"],
  ["Ecovis Malaysia","Audit & accounting","Ecovis global network; serves foreign investors","https://www.ecovis.com.my/","kuala-lumpur@ecovis.com.my","+603 7986 0066","EXSIM Tower, Old Klang Road, KL","https://www.ecovis.com.my/contact-us/kuala-lumpur-hq/","verified"],
  ["Siew Boon Yeong & Associates","Audit & accounting","INAA global association member; intl/MNC clients","https://sby.com.my/","audit@sby.com.my","+603 2693 8837","Jalan Medan Tuanku, KL","https://sby.com.my/contact-us","verified"],
  ["GSK & Associates","Audit & accounting","Audit/tax/company registration for foreign-owned cos","https://www.gskassociates.net/","gunalan@gskassociates.net","+603 6416 0151","MET Corporate Towers, Jln Dutamas 2, KL","https://www.gskassociates.net/","verified"],
  ["3E Accounting Malaysia","Accounting & corporate-secretarial","Incorporation & compliance for foreign investors / MNCs","https://www.3ecpa.com.my/","info@3ecpa.com.my","+603 2387 5300","Menara Bangkok Bank, Jalan Ampang, KL","https://www.3ecpa.com.my/contact-us/","verified"],
  ["Stanton Chase Kuala Lumpur","Executive search / recruitment","Global network, 70+ offices/45 countries; MNC clients","https://www.stantonchase.com/","kualalumpur@stantonchase.com","+60 10 233 5940","Bukit Gasing, PJ","https://www.stantonchase.com/office/executive-search-firm-in-kuala-lumpur-malaysia","verified"],
  ["Pedersen & Partners","Executive search / recruitment","Global executive search; local + international reach","https://pedersenandpartners.com/","Kuala-Lumpur@pedersenandpartners.com","+603 2703 8424","The Intermark, Jalan Tun Razak, KL","https://pedersenandpartners.com/offices/kuala-lumpur/malaysia","verified"],
  ["Monroe Consulting Group Malaysia","Executive search / recruitment","Recruits for multinationals across SE Asia","https://www.monroeconsulting.com/","avinash@monroeconsulting.com.my","+603 2771 0310","Kuala Lumpur","https://www.monroeconsulting.com/executive-search/executive-search-malaysia","verified"],
  ["VERITAS Architects","Architecture & design","Integrated design firm, 7 global offices from KL HQ","https://www.theveritasdesigngroup.com/","enquiry@theveritasdesigngroup.com","+603 4131 6600","Wangsa 118, Wangsa Maju, KL","https://www.theveritasdesigngroup.com/our-offices/","verified"],
  ["RSP Malaysia","Architecture & engineering","Regional RSP group; multidisciplinary regional projects","https://rsp.design/malaysia/","rsp@rsp.design","+603 2273 6636","Plaza Sentral, KL Sentral, KL","https://rsp.design/malaysia/","verified"],
  ["GDP Architects","Architecture","Award-winning KL firm on regional, large-scale projects","https://www.gdparchitects.com/","FORM: https://www.gdparchitects.com/contact/","+603 2095 9500","Bukit Damansara, KL","https://www.gdparchitects.com/contact/","verified"],
  ["Ashton Corporate Services","Corporate-secretarial","Co. secretarial & incorporation for foreign-owned cos","https://www.ashtoncorporate.com/","general@ashtoncorporate.com","+6016 324 0990","Pandan Indah, KL (also Klang/Subang)","https://www.ashtoncorporate.com/","verified"],
  ["MISHU","Corporate-secretarial","Incorporation of foreign cos & rep offices in Malaysia","https://mishu.my/","FORM: https://mishu.my/contact-us/","+60 17 515 9832","Plaza Bukit Jalil, KL","https://mishu.my/","verified"],
  ["Grof Malaysia","Corporate-secretarial","Co. secretary & incorporation for foreign investors","https://grof.co/my/corporate-secretary","FORM: https://grof.co/my/corporate-secretary","+603 6414 1057","Menara Southpoint, Mid Valley, KL","https://grof.co/my/corporate-secretary","verified"],
  ["Secretary Express","Corporate-secretarial","Incorporation & secretarial for foreign-owned Sdn Bhd","https://www.secretaryexpress.com.my/","FORM: https://www.secretaryexpress.com.my/contact-company-secretary-kl/","+603 6276 8897","Segambut, KL","https://www.secretaryexpress.com.my/","verified"],
  ["Impact Communications","PR & communications","Full-service PR/digital; local & international clients","https://impactcommunications.com.my/","FORM: https://impactcommunications.com.my/contact-us/","+60 16 660 0027","Kuala Lumpur","https://impactcommunications.com.my/","verified"],
  ["TQPR Malaysia","PR & communications","Regional PR network; intl/MNC brands since 1995","https://tqpr.com/tqpr-malaysia/","FORM: https://tqpr.com/tqpr-malaysia/","+603 6203 4300","Plaza Damas, Sri Hartamas, KL","https://tqpr.com/tqpr-malaysia/","listed"],
  ["VML Malaysia","Marketing & communications","Global network, 55+ markets; multinational brands","https://www.vml.com/malaysia","FORM: https://www.vml.com/contact","+60 3 7890 3959","Equatorial Plaza, Jln Sultan Ismail, KL","https://www.vml.com/malaysia","listed"],
]

corporate = [
  ["Scicom (MSC) Berhad","BPO / shared services / contact center","Outsourcing & CX services for global brands","https://scicom-intl.com/","business@scicom.com.my","+603 2162 1088","Menara TA One, KLCC, KL","https://scicom-intl.com/contact-us/","verified"],
  ["GP Outsourcing Asia","BPO / EOR / HR outsourcing","HR & workforce solutions for local & intl businesses","https://www.gpoasia.com","recruitment@gpasia.net","+60 17-268 2480","Pusat Perniagaan UOA, Shah Alam","https://www.gpoasia.com/contact","verified"],
  ["TDCX (Malaysia)","BPO / customer experience","Outsourced CX/contact-center for global tech brands","https://www.tdcx.com/my-en/","FORM: https://www.tdcx.com/contact-us/","","Kuala Lumpur","https://www.tdcx.com/my-en/","listed"],
  ["MoneyMatch","Fintech / cross-border payments","International money transfers in 40+ currencies","https://www.moneymatch.co","customer.support@moneymatch.co","+603 3099 3889","Bandar Utama, PJ","https://www.moneymatch.co/contact","verified"],
  ["SleekFlow","Fintech / SaaS (messaging)","Conversational commerce SaaS across multiple markets","https://sleekflow.io","hi@sleekflow.io","+60 3-9212 1090","Naza Tower, Platinum Park, KLCC, KL","https://sleekflow.io/contact","verified"],
  ["Lizard Global","Tech / software & SaaS","Offices in Europe/Asia/Australia; international clients","https://lizard.global","hello@lizard.global","+60 18 356 5702","Bukit Damansara, KL","https://lizard.global/","verified"],
  ["BrioHR","Tech / SaaS (HR software)","Cloud HR platform sold across SE Asian markets","https://briohr.com","FORM: https://briohr.com/contact-us/","","Oval Damansara, TTDI, KL","https://briohr.com/contact-us/","verified"],
  ["LottieFiles","Tech / SaaS (design tooling)","Animation platform with global developer/designer base","https://lottiefiles.com","support@lottiefiles.com","","Menara Etiqa, Bangsar, KL","https://lottiefiles.com/contact-sales","listed"],
  ["Multi-Trans Sdn Bhd","Logistics / freight forwarding","International air & sea freight forwarding","https://www.multi-trans.com.my/","enquiries@multi-trans.com.my","+603 3375 2288","Port Klang + KLIA Cargo Village","https://www.multi-trans.com.my/","verified"],
  ["Quanterm Logistics","Logistics / freight forwarding","Multimodal freight, domestic & intl across Asia-Pacific","https://quanterm.com/","enquiry@quanterm.com","+603 5121 8000","Klang Valley, Selangor","https://quanterm.com/","verified"],
  ["ALFRO Freight Forwarders (M)","Logistics / freight forwarding","Air/sea/land freight via KLIA & Port Klang","https://www.alfrofreight.com/","FORM: https://www.alfrofreight.com/contact/","+603 8778 8491","KLIA Cargo Complex, Sepang + Port Klang","https://www.alfrofreight.com/","verified"],
  ["Pacific Inter-Link (PIL Group)","Trading / commodities exporter","Exports palm oil & commodities to global markets","https://www.pilgroup.com/","info@pilgroup.com","+60 3 4027 1000","Kuala Lumpur","https://www.pilgroup.com/","verified"],
  ["ATA Group Malaysia","Trading / import-export","Trades natural resources & foods internationally","https://www.atagroupsb.com/","info@atagroupsb.com","","Solaris Dutamas, Mont Kiara, KL","https://www.atagroupsb.com/","verified"],
  ["Yang Ma Sdn Bhd","Trading / F&B import-export","Import/export of F&B premix ingredients & OEM products","https://main.yangma.com.my/","info@yangma.com.my","+60 3-2380 0234","Plaza 138, Jalan Ampang, KL","https://main.yangma.com.my/contact-us/","verified"],
  ["Golden Palm Oil Industries","Exporter / palm oil","Refined palm oil packaging, supply & distribution","https://www.goldenpalmoil.com/","alijaya2@yahoo.com","+603 9171 4200","Bandar Tun Razak, Cheras, KL","https://www.goldenpalmoil.com/contact-us","verified"],
  ["KLK OLEO (Kuala Lumpur Kepong)","Exporter / oleochemicals","Global oleochemical producer exporting worldwide","https://www.klkoleo.com/","FORM: https://www.klkoleo.com/contact-us/","+603 7809 8833","Menara KLK, Mutiara Damansara, PJ","https://www.klkoleo.com/contact-us/","verified"],
]

DATA = {
    "HEALTHCARE / MEDICAL": healthcare,
    "HOSPITALITY / TOURISM": hospitality,
    "PROFESSIONAL SERVICES": professional,
    "CORPORATE / EXPORT / TECH": corporate,
}

HEADERS = ["#","Business name","Sub-sector","Why they serve international clients",
           "Website","Email / contact","Phone","Area","Source (verified at)","Confidence","Status","Notes"]

# column widths
WIDTHS = [4,30,28,42,40,38,18,28,42,11,12,26]

thin = Side(style="thin", color="D8DCE4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, fill_color, font_color=WHITE, size=11):
    for c in range(1, len(HEADERS)+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(bold=True, color=font_color, size=size)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = border

def build_kl_sheet(ws):
    # title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    t = ws.cell(row=1, column=1, value="ELEVARE  ·  Kuala Lumpur — B2B outreach targets (businesses serving international clients)")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.font = Font(bold=True, color=GOLD, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    s = ws.cell(row=2, column=1, value="Researched 2026-06-07. 'FORM: <url>' = no public email published — use that contact page. 'verified' = confirmed on official site; 'listed' = from directory/secondary source, double-check before sending.")
    s.fill = PatternFill("solid", fgColor=CREAM)
    s.font = Font(italic=True, color="555555", size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 26

    row = 4
    counter = 1
    for sector, blurb in SECTORS:
        # sector band
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        band = ws.cell(row=row, column=1, value=f"  {sector}   —   {blurb}")
        band.fill = PatternFill("solid", fgColor=ROSE)
        band.font = Font(bold=True, color=WHITE, size=12)
        band.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[row].height = 24
        row += 1
        # column headers
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, NAVY)
        ws.row_dimensions[row].height = 30
        row += 1
        # data
        for rec in DATA[sector]:
            name, subsector, angle, website, email, phone, area, source, conf = rec
            vals = [counter, name, subsector, angle, website, email, phone, area, source, conf, "", ""]
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=10)
            # zebra
            if counter % 2 == 0:
                for i in range(1, len(HEADERS)+1):
                    ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=LIGHT)
            # hyperlinks
            wcell = ws.cell(row=row, column=5)
            if website:
                wcell.hyperlink = website; wcell.font = Font(size=10, color="1155CC", underline="single")
            ecell = ws.cell(row=row, column=6)
            if email.startswith("FORM:"):
                url = email.split("FORM:",1)[1].strip()
                ecell.hyperlink = url
                ecell.font = Font(size=10, color="C2476B", italic=True)
            elif email and "@" in email:
                ecell.hyperlink = f"mailto:{email}"
                ecell.font = Font(size=10, color="1155CC", underline="single")
            scell = ws.cell(row=row, column=9)
            if source:
                scell.hyperlink = source; scell.font = Font(size=9, color="1155CC", underline="single")
            # confidence color
            ccell = ws.cell(row=row, column=10)
            ccell.font = Font(size=10, bold=True,
                              color=("2E7D32" if conf=="verified" else "B26A00"))
            counter += 1
            row += 1
        row += 1  # gap between sectors

    # column widths + freeze
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

# ---- Chiang Mai tracker (retrieved from Debby's Google Sheet 'Elevare_CM_Outreach_Tracker') ----
CM_HEADERS = ["ID","Tier","Sector","Organization","Target Contact (Role)","Contact Name",
              "Email","Phone / Other","Product Pitched","Seats","Est. Monthly Value (USD)",
              "Status","Date Sent","Last Action","Next Action","Notes"]
CM_WIDTHS = [4,5,13,30,28,16,34,17,22,6,14,18,12,12,38,42]
CM_TIER_LABELS = {
    1: ("TIER 1 — HOSPITALS", "Highest ticket ($699/mo Med). Hit these first.", "C0392B"),
    2: ("TIER 2 — HOTELS", "Established L&D budgets, volume Pro deals.", "C9A227"),
    3: ("TIER 3 — TECH & CO-WORKING", "Referral channels, lower ticket but warm.", "2980B9"),
    4: ("TIER 4 — EDUCATION / NGO / MFG", "Case-study fuel, often lower budget.", "27AE60"),
}
CM_DATA = [
  [1,1,"Hospitals","Bangkok Hospital Chiang Mai","Director of Nursing / Nursing Ed Mgr","","info@bangkokhospital-chiangmai.com","+66 52 089 888","Elevare Med",5,"$3,495","Replied — not now","","","Find DoN name on LinkedIn before re-sending","International patient hospital. High motivation among nurses."],
  [2,1,"Hospitals","Chiang Mai Ram Hospital","Director of Nursing","","info@chiangmairam.com","+66 53 920 300","Elevare Med",5,"$3,495","Not contacted","","","Send Med pitch","Large nursing staff, international wing."],
  [3,1,"Hospitals","Lanna Hospital","HR / Nursing Education","","","+66 53 999 777","Elevare Med",3,"$2,097","Not contacted","","","Find email via lanna-hospital.com contact form","Mid-size, growing international patient base."],
  [4,1,"Hospitals","McCormick Hospital","HR Manager","","","+66 53 921 777","Elevare Med",3,"$2,097","Not contacted","","","Find email via mccormick.in.th","Faith-based, supports staff development."],
  [5,1,"Hospitals","Sriphat Medical Center (CMU)","Nursing Education Manager","","sriphat@cmu.ac.th","+66 53 936 900","Elevare Med",5,"$3,495","Not contacted","","","Send Med pitch — CMU-linked","University-linked. Staff pursue further quals."],
  [6,2,"Hospitality","Four Seasons Resort Chiang Mai","Director of Learning & Development","","","+66 53 298 181","Elevare Pro",15,"$4,485","Not contacted","","","Call front desk, ask for L&D Director name","Strong L&D culture across Four Seasons globally."],
  [7,2,"Hospitality","Anantara Chiang Mai Resort","Training Manager","","reservations.chiangmai@anantara.com","+66 53 253 333","Elevare Pro",12,"$3,588","Not contacted","","","Find Training Mgr name on LinkedIn","Large guest-facing team."],
  [8,2,"Hospitality","137 Pillars House Chiang Mai","General Manager","","info@137pillarshouse.com","+66 53 247 788","Elevare Pro",8,"$2,392","Not contacted","","","Pitch GM directly — small property","Boutique, GM controls L&D decisions."],
  [9,2,"Hospitality","Shangri-La Chiang Mai","Director of HR","","slcm@shangri-la.com","+66 53 253 888","Elevare Pro",15,"$4,485","Not contacted","","","Send Pro pitch","Established L&D budget."],
  [10,2,"Hospitality","Le Méridien Chiang Mai","Director of HR","","","+66 53 253 666","Elevare Pro",12,"$3,588","Not contacted","","","Find HR Director on LinkedIn (Marriott)","Marriott corporate L&D framework."],
  [11,2,"Hospitality","Dusit D2 Chiang Mai","HR Manager","","dusitd2.chiangmai@dusit.com","+66 53 999 999","Elevare Pro",10,"$2,990","Not contacted","","","Send Pro pitch","Thai chain, English upskilling priority."],
  [12,2,"Hospitality","Akyra Manor Chiang Mai","GM / HR","","rsvn@theakyra.com","+66 53 216 219","Elevare Pro",6,"$1,794","Not contacted","","","Pitch GM directly — boutique","Boutique, decision-maker reachable."],
  [13,2,"Hospitality","U Nimman Chiang Mai","HR Manager","","reservations.unmm@uhotelsresorts.com","+66 53 005 588","Elevare Pro",8,"$2,392","Not contacted","","","Send Pro pitch","U Hotels chain, modern brand."],
  [14,2,"Hospitality","Centara Riverside Chiang Mai","Training Manager","","crcm@chr.co.th","+66 53 999 999","Elevare Pro",10,"$2,990","Not contacted","","","Verify email then send","Centara has structured L&D."],
  [15,3,"Co-working","4Seas Co-working (HOME BASE)","Community Manager / Founder","","","","Referral partnership",0,"-","Not contacted","","","Have in-person conversation — ask to be recommended training partner","YOUR co-working space. Easiest first conversation."],
  [16,3,"Co-working","Punspace Nimman","Community Manager","","hello@punspace.com","+66 86 911 9293","Referral partnership",0,"-","Not contacted","","","Visit, propose member discount for referrals","Member companies = lead pool."],
  [17,3,"Co-working","CAMP @ Maya","Operations Manager","","","","Referral partnership",0,"-","Not contacted","","","Walk in, ask for ops manager","AIS-run, high foot traffic."],
  [18,3,"Co-working","Alt_ChiangMai","Founder / Manager","","hello@altchiangmai.com","","Referral partnership",0,"-","Not contacted","","","Email + visit","Smaller hub, founder-led."],
  [19,3,"Tech","Pacific Prime (Chiang Mai)","L&D Manager","","careers@pacificprime.com","","Elevare Pro",20,"$5,980","Not contacted","","","Find L&D Manager on LinkedIn","Big CM employer. Insurance — heavy English client work."],
  [20,3,"Tech","Seven Peaks Software","People Operations Manager","","hello@sevenpeakssoftware.com","","Elevare Pro",8,"$2,392","Not contacted","","","Find People Ops on LinkedIn","CM presence; Bangkok HQ."],
  [21,4,"Education","CMU Faculty of Business Administration","Director of Executive Education","","cmubs@cmu.ac.th","+66 53 942 110","Elevare Pro / partnership",0,"-","Not contacted","","","Send partnership pitch","Referral channel for working pros."],
  [22,4,"Education","CMU Intl College of Digital Innovation","Director","","icdi@cmu.ac.th","","Elevare Pro / partnership",0,"-","Not contacted","","","Send partnership pitch","International programmes."],
  [23,4,"Education","Payap University International College","Dean / Director","","inter@payap.ac.th","","Elevare Pro / partnership",0,"-","Not contacted","","","Send partnership pitch","International college, English-medium."],
  [24,4,"NGO","Warm Heart Foundation","Director","","info@warmheartworldwide.org","","Elevare Pro (discounted)",5,"$1,495","Not contacted","","","Offer pilot rate in exchange for case study","Willing to use case studies for marketing."],
  [25,4,"NGO","Urban Light","Executive Director","","info@urban-light.org","","Elevare Pro (discounted)",4,"$1,196","Not contacted","","","Pilot offer + case study","Small team, high English need for fundraising."],
  [26,4,"NGO","The Freedom Story","Executive Director","","info@thefreedomstory.org","","Elevare Pro (discounted)",3,"$897","Not contacted","","","Pilot offer","International donor communication."],
  [27,4,"Manufacturing","IEAT Northern Region (Lamphun)","Tenant Liaison","","","+66 53 581 035","Elevare Pro (volume)",0,"-","Not contacted","","","Call IEAT office, ask for tenant HR referrals","Industrial estate. Many Japanese-owned firms."],
]

def build_chiangmai_sheet(ws):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CM_HEADERS))
    t = ws.cell(row=1, column=1, value="ELEVARE  ·  Chiang Mai — B2B Outreach Tracker   (goal: $1,000 MRR in 30 days → $5,000 in 90)")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.font = Font(bold=True, color=GOLD, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CM_HEADERS))
    s = ws.cell(row=2, column=1, value="Retrieved from your Google Sheet 2026-06-07. Status vocab: Not contacted · Email sent · Follow-up 1 (day 5) · Follow-up 2 (day 12) · Replied — interested/not now · Meeting booked · Proposal sent · Won · Lost · Dead.")
    s.fill = PatternFill("solid", fgColor=CREAM)
    s.font = Font(italic=True, color="555555", size=9)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 26

    row = 4
    for tier in (1, 2, 3, 4):
        label, blurb, color = CM_TIER_LABELS[tier]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(CM_HEADERS))
        band = ws.cell(row=row, column=1, value=f"  {label}   —   {blurb}")
        band.fill = PatternFill("solid", fgColor=color)
        band.font = Font(bold=True, color=WHITE, size=12)
        band.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[row].height = 22
        row += 1
        for i, h in enumerate(CM_HEADERS, start=1):
            ws.cell(row=row, column=i, value=h)
        for c in range(1, len(CM_HEADERS)+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color=WHITE, size=10)
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 28
        row += 1
        for rec in [r for r in CM_DATA if r[1] == tier]:
            for i, v in enumerate(rec, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=10)
            if rec[0] % 2 == 0:
                for i in range(1, len(CM_HEADERS)+1):
                    ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=LIGHT)
            ecell = ws.cell(row=row, column=7)
            em = rec[6]
            if em and "@" in em:
                ecell.hyperlink = f"mailto:{em}"; ecell.font = Font(size=10, color="1155CC", underline="single")
            stcell = ws.cell(row=row, column=12)
            if rec[11].startswith("Replied"):
                stcell.font = Font(size=10, bold=True, color="2E7D32")
            row += 1
        row += 1

    for i, w in enumerate(CM_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

# ---- Email templates (from the CM tracker — reusable for KL too) ----
CM_TEMPLATES = [
    ("Follow-up timing — the rule", "FU1 = 5 business days after first email (no reply). FU2 = 12 days after first email (≈1 week after FU1) — this is the last touch. After FU2: mark 'Replied — not now' or 'Dead' and re-engage in 90 days. Slow sectors (hospitals, MNC procurement): you may stretch FU1 to 7 business days."),
    ("Follow-up 1  (day 5)", "Subject: Re: [original subject]\n\nHi [Name],\n\nQuick nudge in case the earlier note got buried.\n\nThe short version: Elevare runs AI-facilitated English coaching that staff can actually fit around their work, with live tutor sessions for the moments that matter. Pricing is per-seat and structured for L&D budgets.\n\nIf this isn't relevant right now, no worries at all — happy to drop a line in 90 days. If it is, even a 15-minute call would be enough to tell whether it's a fit for [Company].\n\nBest,\nDebby"),
    ("Follow-up 2  (day 12 — break-up)", "Subject: Closing the loop — Elevare for [Company]\n\nHi [Name],\n\nClosing the loop on this one. I don't want to keep cluttering your inbox if the timing isn't right.\n\nIf English training for [Company]'s team becomes a priority later this year, my contact details are below. I'll make a note to reach out again in [Q3/Q4].\n\nWarm regards,\nDebby\nWhatsApp: +66 94 969 0869"),
    ("Pro — first touch", "Subject options: • Hospitality: Business English for [Hotel] guest-facing staff — employer-reimbursable • Tech/Corporate: AI-coached English for [Company]'s working professionals • Generic: A flexible English programme your staff will actually finish\n\nDear [Name / HR Manager],\n\nI'm Debby, founder of Elevare — an AI-facilitated peer-group English coaching platform. I'm reaching out because [Company]'s team interacts regularly with international [guests/clients/partners], and most corporate English training fails them for the same reason: it's scheduled at times staff can't attend, and it teaches grammar instead of the confidence to actually speak.\n\nElevare Pro is built differently:\n• Small peer groups practising real business situations — meetings, emails, presentations, customer conversations\n• AI facilitator that runs sessions and gives each speaker individual feedback\n• Tutor coaching drops from our qualified coaches for the moments AI can't replace\n• Optional WhatsApp practice bot for 15-minute drills between sessions\n\nPricing is $299/month per seat, structured to be employer-reimbursable under L&D budgets. Volume pricing for teams of 10+.\n\nWould you have 20 minutes in the next two weeks for a short call?\n\nBest regards,\nDebby · Founder, Elevare · WhatsApp +66 94 969 0869 · elevaremind.io"),
    ("Med — first touch (hospitals)", "Subject: Helping your nursing staff pass OET — Grade B in 12 weeks\n\nDear [Director of Nursing / HR Director],\n\nI'm Debby, founder of Elevare. I'm writing because many internationally-trained nurses and doctors at hospitals like [Hospital Name] are working toward OET certification.\n\nElevare Med is a 12-week OET preparation programme designed specifically for working clinicians, with a pass-or-keep-coaching guarantee — if a participant doesn't reach OET Grade B, we continue coaching at no extra cost until they do.\n\n• Live small-group coaching with OET-trained tutors\n• AI-facilitated practice between sessions, around shift schedules\n• Full mock exams and exam-month intensive (optional bundle)\n• Clinical scenario practice — handovers, case discussions, written referrals\n\nPricing is $699/month for the 12-week programme. Works as a sponsored development pathway, a retention incentive, or a payroll-funded staff benefit.\n\nWould your nursing leadership be open to a 20-minute conversation?\n\nWarm regards,\nDebby · Founder, Elevare · WhatsApp +66 94 969 0869 · elevaremind.io"),
    ("Proposal email (after 'interested')", "Subject: Elevare proposal for [Company] — [N] seats\n\nHi [Name],\n\nThanks for the conversation. As discussed, here's a costed proposal for [Company]:\n• Programme: Elevare Pro [or Med]\n• Cohort size: [N] seats\n• Monthly investment: $[amount] ($299 × [N], or volume rate for 10+)\n• Start date: [proposed date]\n• Free pilot: one 60-minute sample session for up to 6 staff before sign-off, no commitment\n\nNext: 1) confirm cohort size + start date 2) we run the free pilot 3) sign off and onboard via Whop 4) cohort begins the following Monday.\n\nShall we lock in the pilot session for [week]?\n\nBest,\nDebby"),
]

def build_templates_sheet(ws):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    t = ws.cell(row=1, column=1, value="ELEVARE  ·  Email Templates & Follow-up Cadence  (works for both cities)")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.font = Font(bold=True, color=GOLD, size=14)
    t.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    row = 3
    for title, body in CM_TEMPLATES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        h = ws.cell(row=row, column=1, value=title)
        h.fill = PatternFill("solid", fgColor=ROSE)
        h.font = Font(bold=True, color=WHITE, size=11)
        h.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        b = ws.cell(row=row, column=1, value=body)
        b.font = Font(size=10, color="333333")
        b.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True, indent=1)
        lines = body.count("\n") + max(1, len(body)//90)
        ws.row_dimensions[row].height = max(60, 15 * lines)
        row += 2
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 40
    ws.sheet_view.showGridLines = False

def build_readme(ws):
    rows = [
        ("ELEVARE — Outreach workbook", ""),
        ("", ""),
        ("Purpose", "B2B outreach targets: businesses whose staff deal with international clients and may buy Elevare English training."),
        ("Tabs", "One tab per city so the lists never mix. 'Kuala Lumpur' = new research (72 businesses). 'Chiang Mai' = your existing tracker (27 orgs), merged in from your Google Sheet. 'Email Templates' = pitches + follow-up cadence, usable for both."),
        ("KL sectors", "KL leads grouped into 4 bands: Healthcare/Medical, Hospitality/Tourism, Professional Services, Corporate/Export/Tech."),
        ("CM tiers", "Chiang Mai keeps its original tier system: T1 Hospitals, T2 Hotels, T3 Tech/Co-working, T4 Edu/NGO/Mfg, with Seats + Est. Monthly Value for MRR forecasting."),
        ("Email column", "A real address = published / from your tracker (click to mailto). On KL: 'FORM: <url>' (rose italic) = no public email; click to open their contact page."),
        ("Confidence (KL)", "verified (green) = confirmed on official site 2026-06-07. listed (amber) = directory/secondary source — re-check before sending."),
        ("Follow-up rule", "FU1 at day 5, FU2 (break-up) at day 12, then 90-day re-engage. Full cadence + templates on the Email Templates tab."),
        ("Status column", "Yours to fill: Not contacted / Email sent / Follow-up 1 / Follow-up 2 / Replied / Meeting booked / Proposal sent / Won / Lost / Dead."),
        ("Built", "2026-06-07 · 72 KL businesses (web research) + 27 Chiang Mai orgs (your tracker) · no contact details invented."),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        ca.font = Font(bold=(r==1 or b!=""), size=(15 if r==1 else 11), color=(NAVY if r==1 else "333333"))
        cb.font = Font(size=11, color="333333")
        cb.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 95
    ws.sheet_view.showGridLines = False

wb = Workbook()
ws_readme = wb.active
ws_readme.title = "README"
build_readme(ws_readme)
build_kl_sheet(wb.create_sheet("Kuala Lumpur"))
build_chiangmai_sheet(wb.create_sheet("Chiang Mai"))
build_templates_sheet(wb.create_sheet("Email Templates"))

out = "/home/debby/Desktop/elevare-site/marketing/leads/Elevare-Outreach-Combined.xlsx"
wb.save(out)

total = sum(len(v) for v in DATA.values())
print(f"Saved: {out}")
print(f"Tabs: {wb.sheetnames}")
print(f"KL businesses: {total}")
for s, v in DATA.items():
    ver = sum(1 for r in v if r[8]=="verified")
    print(f"  {s}: {len(v)} ({ver} verified, {len(v)-ver} listed)")
print(f"Chiang Mai orgs: {len(CM_DATA)}")
